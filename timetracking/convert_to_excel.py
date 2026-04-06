#!/usr/bin/env python3
"""
Script to convert overtime_tracking.csv to a properly formatted Excel file.
Requires: pip install openpyxl
"""

import csv
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is not installed.")
    print("Install it with: pip install openpyxl")
    print("Or: python3 -m pip install --user openpyxl")
    sys.exit(1)

def convert_csv_to_excel(csv_file, excel_file):
    """Convert CSV file to formatted Excel file."""
    csv_path = Path(csv_file)
    if not csv_path.exists():
        print(f"Error: CSV file not found: {csv_file}")
        sys.exit(1)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Overtime Tracking'
    
    # Read CSV and write to Excel
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        headers = next(reader)
        
        # Write headers
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Write data rows
        total_hours = 0.0
        evenings_worked = 0
        
        for row_num, row in enumerate(reader, start=2):
            ws.append(row)
            
            # Highlight worked evenings
            worked_col = headers.index('Worked') + 1
            hours_col = headers.index('Hours') + 1
            
            if row[headers.index('Worked')] == 'Yes':
                # Light green background for worked evenings
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    cell.border = thin_border
                
                # Bold hours for worked evenings
                hours_cell = ws.cell(row=row_num, column=hours_col)
                hours_cell.font = Font(bold=True)
                
                total_hours += float(row[headers.index('Hours')])
                evenings_worked += 1
            else:
                # Add borders to non-worked rows
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = thin_border
        
        # Set column widths
        column_widths = {
            'Date': 12,
            'Day': 10,
            'Weekend': 10,
            'Worked': 8,
            'Hours': 8,
            'Commits': 10,
            'Start Time': 10,
            'End Time': 10,
            'Work Summary': 80
        }
        
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            width = column_widths.get(header, 15)
            ws.column_dimensions[col_letter].width = width
        
        # Add summary section
        summary_start = row_num + 3
        ws.cell(row=summary_start, column=1, value='SUMMARY').font = Font(bold=True, size=12)
        summary_start += 1
        
        ws.cell(row=summary_start, column=1, value='Total Hours:')
        ws.cell(row=summary_start, column=2, value=total_hours).font = Font(bold=True)
        summary_start += 1
        
        ws.cell(row=summary_start, column=1, value='Total Evenings Worked:')
        ws.cell(row=summary_start, column=2, value=evenings_worked).font = Font(bold=True)
        summary_start += 1
        
        # Calculate weekday vs weekend
        weekday_hours = 0.0
        weekend_hours = 0.0
        weekday_evenings = 0
        weekend_evenings = 0
        
        for row_num in range(2, ws.max_row + 1):
            if ws.cell(row=row_num, column=headers.index('Worked') + 1).value == 'Yes':
                hours = float(ws.cell(row=row_num, column=headers.index('Hours') + 1).value or 0)
                is_weekend = ws.cell(row=row_num, column=headers.index('Weekend') + 1).value == 'Yes'
                if is_weekend:
                    weekend_hours += hours
                    weekend_evenings += 1
                else:
                    weekday_hours += hours
                    weekday_evenings += 1
        
        ws.cell(row=summary_start, column=1, value='Weekday Hours:')
        ws.cell(row=summary_start, column=2, value=round(weekday_hours, 1)).font = Font(bold=True)
        summary_start += 1
        
        ws.cell(row=summary_start, column=1, value='Weekend Hours:')
        ws.cell(row=summary_start, column=2, value=round(weekend_hours, 1)).font = Font(bold=True)
        summary_start += 1
        
        ws.cell(row=summary_start, column=1, value='Weekday Evenings:')
        ws.cell(row=summary_start, column=2, value=weekday_evenings).font = Font(bold=True)
        summary_start += 1
        
        ws.cell(row=summary_start, column=1, value='Weekend Evenings:')
        ws.cell(row=summary_start, column=2, value=weekend_evenings).font = Font(bold=True)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    # Save Excel file
    wb.save(excel_file)
    print(f"Excel file created successfully: {excel_file}")
    print(f"Total hours: {total_hours:.1f}")
    print(f"Total evenings worked: {evenings_worked}")

if __name__ == '__main__':
    csv_file = Path(__file__).parent / 'overtime_tracking.csv'
    excel_file = Path(__file__).parent / 'overtime_tracking.xlsx'
    
    convert_csv_to_excel(csv_file, excel_file)


